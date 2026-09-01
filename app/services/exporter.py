# services/exporter.py
# Génère les exports Excel (.xlsx) et CSV — extraction à la demande depuis la base
# (la base est la source de vérité, l'Excel est le livrable, section 2 du cadrage).

import csv
import io

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from database import Mouvement, Releve

COULEUR_ENTETE = "1F4E79"
BORDURE = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

NOMS_MOIS = [
    "", "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
    "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre",
]


def _entete(ws, entetes: list[str], hauteur: int = 22):
    for col, texte in enumerate(entetes, start=1):
        c = ws.cell(row=1, column=col, value=texte)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill(start_color=COULEUR_ENTETE, end_color=COULEUR_ENTETE, fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = hauteur


def _feuille_mouvements(wb, titre: str, mouvements: list[Mouvement]):
    ws = wb.create_sheet(titre) if wb.sheetnames != ["Sheet"] else wb.active
    ws.title = titre

    _entete(ws, ["Date", "Libellé", "Débit", "Crédit", "Catégorie"])

    for i, m in enumerate(mouvements, start=2):
        ws.cell(row=i, column=1, value=m.date.strftime("%d/%m/%Y") if m.date else "")
        ws.cell(row=i, column=2, value=m.libelle)
        c_d = ws.cell(row=i, column=3, value=m.debit)
        c_c = ws.cell(row=i, column=4, value=m.credit)
        ws.cell(row=i, column=5, value=m.categorie)

        for cell in (c_d, c_c):
            if cell.value is not None:
                cell.number_format = "#,##0.00"

        if i % 2 == 0:
            for col in range(1, 6):
                ws.cell(row=i, column=col).fill = PatternFill(
                    start_color="EBF3FB", end_color="EBF3FB", fill_type="solid"
                )

        if m.debit:  c_d.font = Font(color="C00000")
        if m.credit: c_c.font = Font(color="375623")

        for col in range(1, 6):
            ws.cell(row=i, column=col).border = BORDURE

    largeurs = {"A": 14, "B": 60, "C": 14, "D": 14, "E": 22}
    for col, larg in largeurs.items():
        ws.column_dimensions[col].width = larg
    ws.freeze_panes = "A2"
    return ws


def _ligne_resume(ws, row, libelle, valeur, gras=False, couleur=None):
    c1 = ws.cell(row=row, column=1, value=libelle)
    c2 = ws.cell(row=row, column=2, value=valeur)
    if gras:
        c1.font = Font(bold=True)
        c2.font = Font(bold=True, color=couleur) if couleur else Font(bold=True)
    if isinstance(valeur, (int, float)):
        c2.number_format = "#,##0.00"


# ── Export d'un relevé unique (section 10 : ré-extraction pour une période traitée) ──

def exporter_releve_excel(societe_nom: str, compte_libelle: str, releve: Releve, mouvements: list[Mouvement]) -> io.BytesIO:
    wb = openpyxl.Workbook()

    ws_resume = wb.active
    ws_resume.title = "Résumé"
    ws_resume.cell(row=1, column=1, value=f"{societe_nom} — {compte_libelle}").font = Font(bold=True, size=13)
    ws_resume.cell(row=2, column=1, value=f"Période : {NOMS_MOIS[releve.mois]} {releve.annee}").font = Font(size=11, color="52525B")

    r = 4
    _ligne_resume(ws_resume, r, "Banque", releve.banque_detectee); r += 1
    _ligne_resume(ws_resume, r, "Solde initial", releve.solde_initial); r += 1
    _ligne_resume(ws_resume, r, "Solde final attendu (saisi)", releve.solde_final_attendu); r += 1
    couleur_calc = "375623" if releve.statut == "ok" else "C00000"
    _ligne_resume(ws_resume, r, "Solde final calculé", releve.solde_final_calcule, gras=True, couleur=couleur_calc); r += 1
    _ligne_resume(ws_resume, r, "Statut", "OK" if releve.statut == "ok" else ("ÉCART" if releve.statut == "ecart" else "En attente")); r += 1
    _ligne_resume(ws_resume, r, "Nombre de mouvements", releve.nb_mouvements); r += 1
    _ligne_resume(ws_resume, r, "Fichier source", releve.nom_fichier); r += 1
    ws_resume.column_dimensions["A"].width = 30
    ws_resume.column_dimensions["B"].width = 30

    _feuille_mouvements(wb, "Mouvements", mouvements)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Export consolidé sur l'exercice (section 4.4 : vérification globale sur 12 mois) ──

def exporter_consolide_excel(
    societe_nom: str,
    compte_libelle: str,
    releves_mouvements: list[tuple[Releve, list[Mouvement]]],
) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws_synthese = wb.active
    ws_synthese.title = "Synthèse exercice"
    ws_synthese.cell(row=1, column=1, value=f"{societe_nom} — {compte_libelle}").font = Font(bold=True, size=13)

    _entete_row = 3
    entetes = ["Mois", "Année", "Solde initial", "Solde final attendu", "Solde final calculé", "Statut", "Nb mouvements"]
    for col, texte in enumerate(entetes, start=1):
        c = ws_synthese.cell(row=_entete_row, column=col, value=texte)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color=COULEUR_ENTETE, end_color=COULEUR_ENTETE, fill_type="solid")

    row = _entete_row + 1
    total_mouvements = 0
    for releve, mouvements in releves_mouvements:
        ws_synthese.cell(row=row, column=1, value=NOMS_MOIS[releve.mois])
        ws_synthese.cell(row=row, column=2, value=releve.annee)
        c_si = ws_synthese.cell(row=row, column=3, value=releve.solde_initial)
        c_sfa = ws_synthese.cell(row=row, column=4, value=releve.solde_final_attendu)
        c_sfc = ws_synthese.cell(row=row, column=5, value=releve.solde_final_calcule)
        c_statut = ws_synthese.cell(row=row, column=6,
                                     value="OK" if releve.statut == "ok" else ("ÉCART" if releve.statut == "ecart" else "En attente"))
        ws_synthese.cell(row=row, column=7, value=releve.nb_mouvements)
        if releve.statut == "ecart":
            c_statut.font = Font(bold=True, color="C00000")
        elif releve.statut == "ok":
            c_statut.font = Font(color="375623")
        for c in (c_si, c_sfa, c_sfc):
            if c.value is not None:
                c.number_format = "#,##0.00"
        total_mouvements += releve.nb_mouvements

        _feuille_mouvements(wb, f"{NOMS_MOIS[releve.mois][:3]} {releve.annee}", mouvements)
        row += 1

    if releves_mouvements:
        premier = releves_mouvements[0][0]
        dernier = releves_mouvements[-1][0]
        r = row + 2
        _ligne_resume(ws_synthese, r, "Solde initial de l'exercice", premier.solde_initial, gras=True); r += 1
        _ligne_resume(ws_synthese, r, "Solde final de l'exercice", dernier.solde_final_calcule, gras=True); r += 1
        ecart_global = round((dernier.solde_final_calcule or 0) - dernier.solde_final_attendu, 2) if dernier.solde_final_calcule is not None else None
        _ligne_resume(ws_synthese, r, "Écart global vs attendu", ecart_global,
                       gras=True, couleur=("C00000" if ecart_global not in (0, 0.0, None) else "375623"))

    for col, larg in {"A": 12, "B": 10, "C": 16, "D": 18, "E": 18, "F": 12, "G": 14}.items():
        ws_synthese.column_dimensions[col].width = larg

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── CSV (export simple des mouvements) ─────────────────────────────────────────

def exporter_csv(mouvements: list[Mouvement]) -> io.StringIO:
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    writer.writerow(["Date", "Libelle", "Debit", "Credit", "Categorie"])
    for m in mouvements:
        writer.writerow([
            m.date.strftime("%d/%m/%Y") if m.date else "",
            m.libelle,
            f"{m.debit:.2f}".replace(".", ",") if m.debit else "",
            f"{m.credit:.2f}".replace(".", ",") if m.credit else "",
            m.categorie or "",
        ])
    buf.seek(0)
    return buf

