# services/fec.py
# Génération de l'export FEC (Fichier des Écritures Comptables) pour le module relevés bancaires
# ET pour le module factures (achats/ventes, évolution du 2026-08-16).
# Écriture à double entrée (une ligne banque + une ligne contrepartie par mouvement, ou plus avec
# TVA) ; exportée en .xlsx dans un format simplifié (Jrl/Date Fact./Piéce/N° Compte/Nom client/
# Débit/Crédit/N° Facture — évolution du 2026-09-02, calée sur le modèle FEC.xlsx fourni).
#
# Numérotation d'écriture (EcritureNum, évolution du 2026-08-16, demande Anis) : un compteur par
# société (Societe.prochain_numero_ecriture), partagé entre le FEC des relevés et celui des
# factures — chaque mouvement ou facture consomme un numéro pour toutes ses lignes (2 ou 3), le
# compteur ne repart jamais de zéro entre deux exports. C'est l'appelant (app.py) qui charge le
# compteur avant génération et persiste la valeur retournée après.

import io
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from database import CompteBancaire, Facture, Mouvement, Releve


def _montant(v):
    """Formate un montant en texte au format FEC (virgule décimale, vide si nul)."""
    if v is None:
        return ""
    return f"{v:.2f}".replace(".", ",")


class FECGenerationError(Exception):
    """Erreur de configuration bloquant la génération du FEC (à afficher tel quel à l'utilisateur)."""


def generer_lignes_fec(
    compte: CompteBancaire,
    mouvements_avec_releve: list[tuple[Mouvement, Releve]],
    compte_defaut_numero: str,
    compte_defaut_libelle: str,
    numero_ecriture_depart: int,
) -> tuple[list[dict], int]:
    """
    Construit les lignes FEC (écriture à double entrée, ou triple avec TVA) pour une liste de mouvements.
    Pour chaque mouvement, la ligne banque (compte comptable + journal du compte bancaire) est suivie de :
    - 1 ligne contrepartie tiers (client/fournisseur/salarié), ou compte comptable ordinaire sans TVA,
    - 2 lignes (HT + TVA) si le compte comptable rattaché a un taux de TVA configuré,
    - 1 ligne compte récurrent par défaut si rien n'est rattaché.
    Retourne (lignes, prochain_numero_ecriture) — chaque mouvement consomme un numéro pour
    l'ensemble de ses lignes.
    """
    lignes = []
    journal_code = compte.journal_comptable or ""
    compte_num_banque = compte.numero_compte_comptable or ""
    compte_lib_banque = compte.libelle
    numero = numero_ecriture_depart

    for m, releve in mouvements_avec_releve:
        date_str = m.date.strftime("%Y%m%d") if m.date else ""
        # mouvement.debit = sortie d'argent du compte bancaire (décaissement)
        # mouvement.credit = entrée d'argent sur le compte bancaire (encaissement)
        montant = m.debit if m.debit else m.credit
        sens_debit = bool(m.debit)   # True = la contrepartie est débitée, False = créditée

        base = {
            "JournalCode": journal_code, "JournalLib": journal_code,
            "EcritureNum": str(numero), "EcritureDate": date_str,
            "CompAuxNum": "", "CompAuxLib": "",
            "PieceRef": "", "PieceDate": date_str,
            "EcritureLib": m.libelle,
            "ValidDate": "", "Montantdevise": "", "Idevise": "",
            "DateRglt": "", "ModeRglt": "", "NatOp": "", "IdClient": "",
        }

        def _ligne(compte_num, compte_lib, montant_ligne, debit):
            return {
                **base,
                "CompteNum": compte_num or "", "CompteLib": compte_lib or "",
                "Debit": _montant(montant_ligne if debit else None),
                "Credit": _montant(montant_ligne if not debit else None),
                "DebitDevise": "", "CreditDevise": "",
            }

        # Ligne banque : mouvement au débit -> banque créditée, et inversement
        lignes.append(_ligne(compte_num_banque, compte_lib_banque, montant, not sens_debit))

        tiers = m.tiers if (m.tiers_id and m.tiers) else None
        if tiers and m.taux_tva_id and m.taux_tva:
            taux = m.taux_tva.taux
            compte_tva_num = m.taux_tva.numero_compte_tva
            if not compte_tva_num:
                raise FECGenerationError(
                    f"Le taux de TVA {taux:g}% choisi sur le mouvement « {m.libelle} » n'a pas de compte "
                    "de TVA renseigné — complétez-le dans Paramètres avant d'exporter le FEC."
                )
            montant_ht = round(montant / (1 + taux / 100), 2)
            montant_tva = round(montant - montant_ht, 2)
            lignes.append(_ligne(tiers.numero_compte, tiers.intitule, montant_ht, sens_debit))
            lignes.append(_ligne(compte_tva_num, m.taux_tva.intitule, montant_tva, sens_debit))
        elif tiers:
            lignes.append(_ligne(tiers.numero_compte, tiers.intitule, montant, sens_debit))
        else:
            lignes.append(_ligne(compte_defaut_numero, compte_defaut_libelle, montant, sens_debit))

        numero += 1

    return lignes, numero


def generer_lignes_fec_factures(
    factures: list[Facture],
    journal_achats: str,
    journal_ventes: str,
    compte_defaut_numero: str,
    compte_defaut_libelle: str,
    numero_ecriture_depart: int,
) -> tuple[list[dict], int]:
    """
    Construit les lignes FEC pour une liste de factures (achats et/ou ventes mélangés — le
    JournalCode distingue). Écriture à double entrée, ou triple si un taux de TVA est rattaché :
    - Achat  : Débit compte de charge (HT) [+ Débit compte TVA déductible]  / Crédit fournisseur (TTC)
    - Vente  : Débit client (TTC) / Crédit compte de produit (HT) [+ Crédit compte TVA collectée]
    Sans compte de charge/produit rattaché, repli sur le compte récurrent par défaut (même logique
    que le module relevés). Le tiers (fournisseur/client) est en revanche obligatoire — une facture
    sans tiers rattaché bloque l'export, plutôt que de la rattacher silencieusement à un compte
    générique. Retourne (lignes, prochain_numero_ecriture).
    """
    lignes = []
    numero = numero_ecriture_depart

    for f in factures:
        ref = f.numero or f.fichier_source

        # Le tiers (fournisseur/client) est prioritaire ; à défaut, le compte récurrent choisi en
        # remplacement (ex. compte d'attente, tant que le vrai tiers n'est pas encore identifié) —
        # évolution du 2026-08-16. Ni l'un ni l'autre -> bloque, jamais de repli silencieux ici.
        contrepartie = f.tiers or f.compte_recurrent
        if not contrepartie:
            raise FECGenerationError(
                f"La facture « {ref} » n'a pas de tiers (fournisseur/client) ni de compte d'attente "
                "rattaché — complétez-la avant d'exporter le FEC."
            )
        if f.montant_ttc is None:
            raise FECGenerationError(f"La facture « {ref} » n'a pas de montant TTC renseigné.")

        journal_code = journal_achats if f.type == "achat" else journal_ventes
        date_str = f.date.strftime("%Y%m%d") if f.date else ""
        montant_ttc = f.montant_ttc
        tiers = contrepartie
        compte = f.compte
        compte_num = (compte.numero_compte if compte else None) or compte_defaut_numero
        compte_lib = (compte.intitule if compte else None) or compte_defaut_libelle

        base = {
            "JournalCode": journal_code, "JournalLib": journal_code,
            "EcritureNum": str(numero), "EcritureDate": date_str,
            "CompAuxNum": "", "CompAuxLib": "",
            "PieceRef": f.numero or "", "PieceDate": date_str,
            "EcritureLib": f.nature_prestation or tiers.intitule,
            "ValidDate": "", "Montantdevise": "", "Idevise": "",
            "DateRglt": "", "ModeRglt": "", "NatOp": "", "IdClient": "",
        }

        def _ligne(compte_num, compte_lib, montant_ligne, debit):
            return {
                **base,
                "CompteNum": compte_num or "", "CompteLib": compte_lib or "",
                "Debit": _montant(montant_ligne if debit else None),
                "Credit": _montant(montant_ligne if not debit else None),
                "DebitDevise": "", "CreditDevise": "",
            }

        avec_tva = bool(f.taux_tva_id and f.taux_tva and f.montant_ht is not None and f.montant_tva)
        compte_tva_num = None
        if avec_tva:
            compte_tva_num = f.taux_tva.numero_compte_tva
            if not compte_tva_num:
                raise FECGenerationError(
                    f"Le taux de TVA {f.taux_tva.taux:g}% choisi sur la facture « {ref} » n'a pas "
                    "de compte de TVA renseigné — complétez-le dans Paramètres avant d'exporter le FEC."
                )

        if f.type == "achat":
            if avec_tva:
                lignes.append(_ligne(compte_num, compte_lib, f.montant_ht, True))
                lignes.append(_ligne(compte_tva_num, f.taux_tva.intitule, f.montant_tva, True))
            else:
                lignes.append(_ligne(compte_num, compte_lib, montant_ttc, True))
            lignes.append(_ligne(tiers.numero_compte, tiers.intitule, montant_ttc, False))
        else:  # vente
            lignes.append(_ligne(tiers.numero_compte, tiers.intitule, montant_ttc, True))
            if avec_tva:
                lignes.append(_ligne(compte_num, compte_lib, f.montant_ht, False))
                lignes.append(_ligne(compte_tva_num, f.taux_tva.intitule, f.montant_tva, False))
            else:
                lignes.append(_ligne(compte_num, compte_lib, montant_ttc, False))

        numero += 1

    return lignes, numero


def _defaire_montant(texte: str):
    """Inverse de _montant() : '1234,56' -> 1234.56 (float), '' -> None."""
    if not texte:
        return None
    return float(texte.replace(",", "."))


ENTETES_FEC_SIMPLIFIE = ["Jrl", "Date Fact.", "Piéce", "N° Compte", "Nom client", "Débit", "Crédit", "N° Facture"]


def exporter_fec_xlsx(lignes: list[dict]) -> io.BytesIO:
    """Écrit les lignes FEC (mêmes lignes que exporter_fec_texte, écriture à double entrée)
    au format Excel simplifié demandé par l'utilisateur (calé sur son modèle FEC.xlsx) :
    une ligne par écriture élémentaire (banque, puis contrepartie(s)), colonnes réduites à
    Jrl/Date Fact./Piéce/N° Compte/Nom client/Débit/Crédit/N° Facture — mêmes valeurs que
    le FEC complet (JournalCode/EcritureDate/EcritureNum/CompteNum/CompteLib/Debit/Credit/
    PieceRef), juste renommées et sans les colonnes annexes (journal libellé, comptes
    auxiliaires, devise...) absentes de ce modèle."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FEC"

    for col, texte in enumerate(ENTETES_FEC_SIMPLIFIE, start=1):
        c = ws.cell(row=1, column=col, value=texte)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    for i, ligne in enumerate(lignes, start=2):
        date_fact = None
        if ligne.get("EcritureDate"):
            try:
                date_fact = datetime.strptime(ligne["EcritureDate"], "%Y%m%d").date()
            except ValueError:
                date_fact = None
        ws.cell(row=i, column=1, value=ligne.get("JournalCode") or "")
        c_date = ws.cell(row=i, column=2, value=date_fact)
        if date_fact:
            c_date.number_format = "DD/MM/YYYY"
        ws.cell(row=i, column=3, value=ligne.get("EcritureNum") or "")
        ws.cell(row=i, column=4, value=ligne.get("CompteNum") or "")
        ws.cell(row=i, column=5, value=ligne.get("CompteLib") or "")
        ws.cell(row=i, column=6, value=_defaire_montant(ligne.get("Debit")))
        ws.cell(row=i, column=7, value=_defaire_montant(ligne.get("Credit")))
        ws.cell(row=i, column=8, value=ligne.get("PieceRef") or "")

    for col, largeur in zip("ABCDEFGH", [8, 12, 8, 14, 30, 12, 12, 14]):
        ws.column_dimensions[col].width = largeur

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

